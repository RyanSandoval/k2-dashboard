#!/usr/bin/env python3
"""requests_sync.py — ingest Ryan's M365 Copilot "unresolved requests review"
workbook into DATA.requests on RyanSandoval/k2-data.

Why this exists: K2 has no Microsoft 365 access, so every request that arrives
by Outlook, Teams, or meeting notes is invisible to it. Copilot can see those
surfaces. Ryan exports the review, drops the xlsx, and this script makes the
open asks first-class objects in the dashboard.

Design notes:
  - Stable id = sha256(requester + requested action)[:12] so re-drops of the
    same ask keep their identity, and Ryan's resolve/dismiss survives.
  - User state (resolved / dismissed / note) is read off the LIVE payload and
    re-applied. The spreadsheet never clobbers a human decision.
  - An item present last drop but absent this drop is NOT deleted; it is marked
    goneFromSource so a Copilot coverage gap can't silently erase work.
  - All writes go through k2_data_set.py (the only safe writer).

Usage:
  requests_sync.py --xlsx FILE.xlsx --dry-run
  requests_sync.py --xlsx FILE.xlsx --write
"""
import argparse
import base64
import datetime
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile

REPO = "RyanSandoval/k2-data"
WRITER = os.path.expanduser("~/.openclaw/workspace/scripts/k2_data_set.py")
SHEETS = ("Overdue", "Due Soon", "No Due Date")
BUCKET_KEY = {"Overdue": "overdue", "Due Soon": "due_soon", "No Due Date": "no_due_date"}


def die(msg):
    print("requests_sync ERROR: " + msg, file=sys.stderr)
    sys.exit(1)


def clean(v, limit=None):
    if v is None:
        return ""
    s = re.sub(r"\s+", " ", str(v)).strip()
    if s.lower() in ("none", "nan"):
        return ""
    if limit and len(s) > limit:
        s = s[: limit - 1].rstrip() + "…"
    return s


def iso_date(v):
    s = clean(v)
    if not s:
        return ""
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else ""


def days_since(d, today):
    if not d:
        return None
    try:
        return (today - datetime.date.fromisoformat(d)).days
    except ValueError:
        return None


def first_url(v):
    m = re.search(r"https?://\S+", clean(v))
    return m.group(0).rstrip(".,;") if m else ""


def parse(xlsx, today):
    try:
        import openpyxl
    except ImportError:
        die("openpyxl not installed (pip3 install openpyxl)")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    generated = ""
    if "Summary" in wb.sheetnames:
        for row in wb["Summary"].iter_rows(values_only=True):
            if row and clean(row[0]).lower() == "generated":
                generated = iso_date(row[1])

    rows = []
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        recs = list(wb[sheet].iter_rows(values_only=True))
        if not recs:
            continue
        hdr = [clean(h) for h in recs[0]]
        for raw in recs[1:]:
            if not any(c is not None and clean(c) for c in raw):
                continue
            r = dict(zip(hdr, raw))
            requester = clean(r.get("Requester"), 80)
            ask = clean(r.get("Requested Action"), 400)
            if not ask:
                continue
            requested = iso_date(r.get("Date Requested"))
            due = iso_date(r.get("Due Date"))
            age = days_since(requested, today)
            overdue_by = days_since(due, today)
            rid = "req_" + hashlib.sha256(
                (requester.lower() + "|" + ask.lower()).encode("utf-8")
            ).hexdigest()[:12]
            jira = sorted(set(re.findall(
                r"\b(?:MW|DES|CMS|DO)-\d+\b",
                ask + " " + clean(r.get("Supporting Reference(s)")))))
            rows.append({
                "id": rid,
                "bucket": BUCKET_KEY.get(sheet, sheet),
                "requester": requester,
                "requesterEmail": clean(r.get("Requester Email"), 120),
                "channel": clean(r.get("Source"), 40),
                "ask": ask,
                "requestedDate": requested,
                "dueDate": due,
                "dueType": clean(r.get("Due Date Type"), 20),
                "ageDays": age,
                "overdueDays": (overdue_by if (overdue_by is not None and overdue_by > 0) else 0),
                "whyUnresolved": clean(r.get("Why Unresolved"), 300),
                "nextAction": clean(r.get("Recommended Next Action"), 300),
                "reference": clean(r.get("Supporting Reference(s)"), 140),
                "link": first_url(r.get("Microsoft 365 Link(s)")),
                "resolved": False,
                "resolvedAt": None,
                "dismissed": False,
                "goneFromSource": False,
            })
    # id collision guard: identical requester+ask in two buckets
    seen = {}
    for r in rows:
        if r["id"] in seen:
            r["id"] = r["id"] + "b"
        seen[r["id"]] = True
    return rows, generated


def fetch_live():
    sha = subprocess.run(
        ["gh", "api", "repos/%s/contents/data.json" % REPO, "--jq", ".sha"],
        capture_output=True, text=True)
    if sha.returncode != 0:
        die("could not read data.json sha: " + sha.stderr.strip())
    blob = subprocess.run(
        ["gh", "api", "repos/%s/git/blobs/%s" % (REPO, sha.stdout.strip()), "--jq", ".content"],
        capture_output=True, text=True)
    if blob.returncode != 0:
        die("could not read data.json blob: " + blob.stderr.strip())
    return json.loads(base64.b64decode(blob.stdout))


def merge(new_rows, live):
    """Carry Ryan's decisions forward; never let the sheet undo a human."""
    prior = {r.get("id"): r for r in (live.get("requests") or []) if isinstance(r, dict)}
    out = []
    for r in new_rows:
        old = prior.get(r["id"])
        if old:
            for field in ("resolved", "resolvedAt", "dismissed", "note", "firstSeenAt"):
                if field in old and old[field] not in (None, "", False):
                    r[field] = old[field]
        r.setdefault("firstSeenAt", datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z")
        out.append(r)

    incoming = {r["id"] for r in new_rows}
    carried = []
    for rid, old in prior.items():
        if rid in incoming:
            continue
        old = dict(old)
        old["goneFromSource"] = True
        carried.append(old)
    out.extend(carried)

    added = [r for r in new_rows if r["id"] not in prior]
    return out, added, carried


def sort_key(r):
    # Unresolved first, then most overdue, then oldest ask.
    return (
        1 if (r.get("resolved") or r.get("dismissed")) else 0,
        -(r.get("overdueDays") or 0),
        -(r.get("ageDays") or 0),
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--print-ids", action="store_true")
    ap.add_argument("--today", default=None, help="YYYY-MM-DD override for age math")
    a = ap.parse_args()
    if not a.write and not a.dry_run and not a.print_ids:
        a.dry_run = True

    today = datetime.date.fromisoformat(a.today) if a.today else datetime.date.today()
    rows, generated = parse(a.xlsx, today)
    if not rows:
        die("no request rows parsed from %s" % a.xlsx)

    if a.print_ids:
        for r in sorted(rows, key=lambda x: x["id"]):
            print(r["id"])
        return

    live = fetch_live()
    pre_keys = len(live)
    merged, added, carried = merge(rows, live)
    merged.sort(key=sort_key)

    open_rows = [r for r in merged if not r.get("resolved") and not r.get("dismissed") and not r.get("goneFromSource")]
    meta = {
        "count": len(merged),
        "open": len(open_rows),
        "overdue": len([r for r in open_rows if (r.get("overdueDays") or 0) > 0]),
        "sourceGenerated": generated,
        "sourceFile": os.path.basename(a.xlsx),
        "ingestedAt": datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "newThisDrop": len(added),
        "goneFromSource": len(carried),
    }

    # Variance report — this is the deliverable, the store is plumbing.
    print("=== Requests ingest (%s) ===" % (generated or "no date"))
    print("parsed        : %d" % len(rows))
    print("new this drop : %d" % len(added))
    print("gone from src : %d" % len(carried))
    print("open          : %d  (overdue %d)" % (meta["open"], meta["overdue"]))
    top = [r for r in open_rows if (r.get("overdueDays") or 0) > 0][:5]
    if top:
        print("most overdue  :")
        for r in top:
            print("   %4dd  %-22s %s" % (r["overdueDays"], r["requester"][:22], r["ask"][:60]))
    print("pre-write keys: %d" % pre_keys)
    print("rows=%d" % len(rows))

    if not a.write:
        return

    with tempfile.TemporaryDirectory() as td:
        rp = os.path.join(td, "requests.json")
        mp = os.path.join(td, "requestsMeta.json")
        json.dump(merged, open(rp, "w"))
        json.dump(meta, open(mp, "w"))
        cmd = [sys.executable, WRITER,
               "--set", "requests=" + rp,
               "--set", "requestsMeta=" + mp,
               "--message", "Requests ingest: %d open (%d overdue) from Copilot review %s"
                            % (meta["open"], meta["overdue"], generated or "n/a")]
        res = subprocess.run(cmd)
        if res.returncode != 0:
            die("k2_data_set.py refused the write (exit %d)" % res.returncode)
    print("WROTE ok")


if __name__ == "__main__":
    main()
